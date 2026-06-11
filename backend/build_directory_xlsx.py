# -*- coding: utf-8 -*-
"""One Excel workbook with 4 sheets (Departments / Teams / Clients / Employees),
using the NEW accurate per-activity attribution (task-ID link + project-name
fallback) + name normalization. No double-counting. Output -> app/Org_Directory.xlsx
"""
import db
import pandas as pd

OUT = "../Org_Directory_v2.xlsx"
OPS = ["Titans", "Syndicate", "Synergy", "Alliance", "Falcons", "Mavericks", "Bravix"]
INTERNAL_KW = ["emailing", "training", "nb tasks", "admin-operational", "admin operational",
               "introductory", "operational works", "company maintenance", "master project",
               "new initiatives", "initiatives", "tracker", "audit & reporting", "employee process",
               "accounting - finovate", "marketing -ledgerlabs", "marketing - ledgerlabs",
               "& management", "maintenance"]


def norm_team(raw):
    s = (raw or "").strip(); low = s.lower()
    if not s or low == "(unknown)":
        return "Unassigned"
    if "archived" in low:
        return "Archived Projects"
    if "operations" in low and "ledger" not in low:
        for t in OPS:
            if t.lower() in low:
                return "Operations - " + t
        return "Operations (other)"
    if "ledger labs" in low or low.startswith("ll:"):
        return "Ledger Labs (Internal)"
    if "marketing" in low:
        return "Marketing"
    if "training" in low:
        return "Training"
    if "admin" in low:
        return "Admin"
    if "onboarding" in low:
        return "Onboarding"
    if "tax" in low or "client tracking" in low:
        return "Tax"
    if low.startswith("hr"):
        return "HR"
    if "audit" in low:
        return "Audit"
    if any(k in low for k in ["developer", "odoo", "it ai", "r&d", "automation", "config"]):
        return "IT / R&D"
    if "praduman" in low:
        return "Praduman Singh"
    if "personal" in low:
        return "Personal"
    if any(k in low for k in ["test supabase", "sales", "indian"]):
        return "Unassigned"
    return s


def dept_of(team):
    return "Operations" if team.startswith("Operations") else team


def client_kind(client):
    c = (client or "").lower()
    if any(k in c for k in INTERNAL_KW):
        return "Internal"
    cz = c.replace(" ", "")
    if "(f)" in cz:
        return "Fixed"
    if "(h)" in cz:
        return "Hourly"
    return "Project"


# ---------- per (employee, date, team, client) grain ----------
g = db.q("""
  SELECT a.user_id::text uid, a.user_name nm, a.date::text d,
         CASE WHEN c.task_id IS NOT NULL AND coalesce(c.space_name,'')<>'' THEN c.space_name
              ELSE coalesce(nullif(split_part(pr.name,' / ',1),''),'(unknown)') END team_raw,
         CASE WHEN c.task_id IS NOT NULL AND coalesce(c.folder_name,'')<>'' THEN c.folder_name
              WHEN pr.name LIKE '%/%' THEN trim(split_part(pr.name,' / ',2))
              ELSE coalesce(nullif(pr.name,''),'(no client)') END client_raw,
         sum(a.tracked)/3600.0 h,
         sum(CASE WHEN trim(coalesce(ht.summary,'')) ~* '(^|[^a-z0-9])nb([^a-z0-9]|$)'
                    OR trim(coalesce(pr.name,'')) ~* '(^|[^a-z0-9])nb([^a-z0-9]|$)'
                  THEN a.tracked ELSE 0 END)/3600.0 nb,
         sum(CASE WHEN c.task_id IS NOT NULL THEN a.tracked ELSE 0 END)/3600.0 matched_h
  FROM hubstaff_activities a
  LEFT JOIN hubstaff_tasks ht ON ht.id=a.task_id
  LEFT JOIN clickup_tasks c ON c.task_id=ht.remote_id
  LEFT JOIN hubstaff_projects pr ON pr.id=a.project_id
  WHERE coalesce(a.tracked,0)>0
  GROUP BY 1,2,3,4,5
""")
g["team"] = g["team_raw"].map(norm_team)
g["dept"] = g["team"].map(dept_of)
g["client"] = g["client_raw"].fillna("(no client)").replace("", "(no client)")
g["kind"] = g["client"].map(client_kind)
g["bill"] = (g["h"] - g["nb"]).clip(lower=0)
latest = g["d"].max()
cut30 = (pd.to_datetime(latest) - pd.Timedelta(days=30)).strftime("%Y-%m-%d")
cut90 = (pd.to_datetime(latest) - pd.Timedelta(days=90)).strftime("%Y-%m-%d")


def status(last, h30, h90):
    if h30 > 0:
        return "ACTIVE"
    if h90 > 0:
        return "IDLE (no work 30d)"
    return "INACTIVE"


def win(df, col):
    last = df.groupby(col)["d"].max()
    h30 = df[df["d"] >= cut30].groupby(col)["h"].sum()
    h90 = df[df["d"] >= cut90].groupby(col)["h"].sum()
    return last, h30, h90


total = g["h"].sum()

def linked_pct(df, col):
    m = df.groupby(col)["matched_h"].sum()
    h = df.groupby(col)["h"].sum()
    return {k: round(m.get(k, 0) / h.get(k, 1) * 100, 0) for k in h.index}


# ---------- DEPARTMENTS ----------
dl, d30, d90 = win(g, "dept")
dep_lp = linked_pct(g, "dept")
dep = (g.groupby("dept").agg(Hours=("h", "sum"), Teams=("team", "nunique"),
                            Employees=("uid", "nunique"), Clients=("client", "nunique")).reset_index())
dep["Status"] = dep["dept"].map(lambda x: status(dl.get(x), d30.get(x, 0), d90.get(x, 0)))
dep["Hours_last_30d"] = dep["dept"].map(lambda x: round(d30.get(x, 0), 1))
dep["Pct_of_total"] = (dep["Hours"] / total * 100).round(1)
dep["Linked"] = dep["dept"].map(lambda x: dep_lp.get(x, 0))
dep["Hours"] = dep["Hours"].round(1)
dep = dep.sort_values("Hours", ascending=False)
dep = dep[["dept", "Status", "Hours", "Pct_of_total", "Hours_last_30d", "Teams", "Employees", "Clients", "Linked"]]
dep.columns = ["Department", "Status", "Total Hours", "% of total", "Hours last 30d", "Teams", "Employees", "Clients", "ClickUp-linked %"]

# ---------- TEAMS ----------
tl, t30, t90 = win(g, "team")
tm = (g.groupby("team").agg(Department=("dept", "first"), Hours=("h", "sum"),
                            Employees=("uid", "nunique"), Clients=("client", "nunique")).reset_index())
tm["Status"] = tm["team"].map(lambda x: status(tl.get(x), t30.get(x, 0), t90.get(x, 0)))
tm["Hours_30d"] = tm["team"].map(lambda x: round(t30.get(x, 0), 1))
tm["Last_worked"] = tm["team"].map(lambda x: tl.get(x))
tm_lp = linked_pct(g, "team")
tm["Linked"] = tm["team"].map(lambda x: tm_lp.get(x, 0))
tm["Hours"] = tm["Hours"].round(1)
tm = tm.sort_values("Hours", ascending=False)
tm = tm[["team", "Department", "Status", "Hours", "Hours_30d", "Last_worked", "Employees", "Clients", "Linked"]]
tm.columns = ["Team", "Department", "Status", "Total Hours", "Hours last 30d", "Last worked", "Employees", "Clients", "ClickUp-linked %"]

# ---------- CLIENTS ----------
cl, c30, c90 = win(g, "client")
# dominant team/dept per client (most hours)
dom = (g.groupby(["client", "team"])["h"].sum().reset_index()
       .sort_values("h", ascending=False).drop_duplicates("client").set_index("client"))
cli = (g.groupby("client").agg(Hours=("h", "sum"), Billable=("bill", "sum"),
                               Employees=("uid", "nunique")).reset_index())
cli["Kind"] = cli["client"].map(client_kind)
cli["Team"] = cli["client"].map(lambda x: dom.loc[x, "team"] if x in dom.index else "")
cli["Department"] = cli["Team"].map(lambda x: dept_of(x) if x else "")
cli["Status"] = cli["client"].map(lambda x: status(cl.get(x), c30.get(x, 0), c90.get(x, 0)))
cli["Hours_30d"] = cli["client"].map(lambda x: round(c30.get(x, 0), 1))
cli["Last_worked"] = cli["client"].map(lambda x: cl.get(x))
cli_lp = linked_pct(g, "client")
cli["Linked"] = cli["client"].map(lambda x: cli_lp.get(x, 0))
cli["Hours"] = cli["Hours"].round(1); cli["Billable"] = cli["Billable"].round(1)
cli = cli.sort_values("Hours", ascending=False)
cli = cli[["client", "Kind", "Status", "Department", "Team", "Hours", "Billable", "Hours_30d", "Last_worked", "Employees", "Linked"]]
cli.columns = ["Client", "Kind", "Status", "Department", "Team", "Total Hours", "Billable Hours", "Hours last 30d", "Last worked", "Employees", "ClickUp-linked %"]

# ---------- EMPLOYEES ----------
el = g.groupby("uid")["d"].max()
e30 = g[g["d"] >= cut30].groupby("uid")["h"].sum()
e90 = g[g["d"] >= cut90].groupby("uid")["h"].sum()
prim_t = g.groupby(["uid", "team"])["h"].sum().reset_index().sort_values("h", ascending=False).drop_duplicates("uid").set_index("uid")["team"]
emp = (g.groupby("uid").agg(Employee=("nm", "first"), Hours=("h", "sum"), Billable=("bill", "sum"),
                            NB=("nb", "sum"), Teams=("team", "nunique"), Clients=("client", "nunique"),
                            Active_days=("d", "nunique")).reset_index())
emp["Status"] = emp["uid"].map(lambda x: status(el.get(x), e30.get(x, 0), e90.get(x, 0)))
emp["Last_worked"] = emp["uid"].map(lambda x: el.get(x))
emp["Primary_Team"] = emp["uid"].map(lambda x: prim_t.get(x, ""))
emp["Primary_Department"] = emp["Primary_Team"].map(lambda x: dept_of(x) if x else "")
emp_lp = linked_pct(g, "uid")
emp["Linked"] = emp["uid"].map(lambda x: emp_lp.get(x, 0))
for c in ["Hours", "Billable", "NB"]:
    emp[c] = emp[c].round(1)
emp = emp.sort_values("Hours", ascending=False)
emp = emp[["Employee", "Status", "Last_worked", "Primary_Department", "Primary_Team",
           "Teams", "Clients", "Hours", "Billable", "NB", "Active_days", "Linked"]]
emp.columns = ["Employee", "Status", "Last worked", "Primary Department", "Primary Team",
               "# Teams", "# Clients", "Total Hours", "Billable Hours", "NonBillable Hours", "Active days", "ClickUp-linked %"]

# ---------- Read Me / Sources sheet ----------
src = pd.DataFrame([
    ["Department", "ClickUp (primary) / Hubstaff (fallback)", "clickup_tasks.space_name  OR  hubstaff_projects.name prefix",
     "ClickUp space ka ' - ' se pehla hissa; agar task link na ho to Hubstaff project naam se"],
    ["Team", "ClickUp (primary) / Hubstaff (fallback)", "clickup_tasks.space_name  OR  hubstaff_projects.name prefix",
     "Poora ClickUp space; Operations ke sub-team (Titans, Bravix...) alag"],
    ["Client", "ClickUp (primary) / Hubstaff (fallback)", "clickup_tasks.folder_name  OR  hubstaff project ' / ' ke baad",
     "ClickUp folder = client; warna Hubstaff project naam ka client hissa"],
    ["Kind (client type)", "ClickUp folder naam", "folder_name marker",
     "(F)=Fixed, (H)=Hourly, internal kaam=Internal, warna Project"],
    ["Employee", "Hubstaff", "hubstaff_activities.user_name", "Hubstaff naam, ClickUp assignee se match"],
    ["Total / Billable / NB Hours", "Hubstaff", "hubstaff_activities.tracked", "seconds / 3600; NB = task/project naam me 'NB' marker"],
    ["Active days / Last worked", "Hubstaff", "hubstaff_activities.date", "jin dino time track hua"],
    ["Status", "Hubstaff (derived)", "max(date) vs latest", "ACTIVE=last 30d, IDLE=last 90d, INACTIVE=90d+ koi time nahi"],
    ["ClickUp-linked %", "Hubstaff x ClickUp", "remote_id = task_id match",
     "us row ke kitne ghante ClickUp task se PAKKA jude (baaki Hubstaff project naam se anumaan)"],
    ["", "", "", ""],
    ["THE LINK", "Hubstaff -> ClickUp", "hubstaff_tasks.remote_id = clickup_tasks.task_id",
     "integration 189908; match par task naam 99.97% same; ek-tarfa link"],
    ["Overall match", "", "", "~53% time ClickUp task se exact linked (2026: ~62%); baaki project-level/unmatched"],
], columns=["Column / Field", "Source System", "Table / Field", "Rule / Note"])

# ---------- write workbook with formatting ----------
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    src.to_excel(xw, sheet_name="Read Me - Sources", index=False)
    dep.to_excel(xw, sheet_name="Departments", index=False)
    tm.to_excel(xw, sheet_name="Teams", index=False)
    cli.to_excel(xw, sheet_name="Clients", index=False)
    emp.to_excel(xw, sheet_name="Employees", index=False)
    wb = xw.book
    head_fill = PatternFill("solid", fgColor="0F172A")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="E2E8F0")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = head_fill; cell.font = head_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        # column widths
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 11), 46)
        # zebra + borders
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for c in row:
                c.border = Border(bottom=thin)
                if i % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="F8FAFC")

print("WROTE", OUT)
print(f"Departments={len(dep)}  Teams={len(tm)}  Clients={len(cli)}  Employees={len(emp)}")
print(f"check totals -> dept={dep['Total Hours'].sum():,.0f}h  team={tm['Total Hours'].sum():,.0f}h  "
      f"emp={emp['Total Hours'].sum():,.0f}h  (actual {total:,.0f}h)")
