# -*- coding: utf-8 -*-
"""Reconcile dashboard (Hubstaff) employees against HR files:
 - Job Details (active roster) + Relieved Employees (left).
Output -> app/Employee_Status.xlsx with sheets:
 Active / Relieved / External-Client / Unknown-Staff + Summary."""
import re
import difflib
import pandas as pd
import db

JOB = r"C:\Users\ADMIN\Downloads\Job Details - Finovate Consulting Pvt Ltd (1).xlsx"
REL = r"C:\Users\ADMIN\Downloads\Relieved Employees - Finovate Consulting Pvt Ltd.xlsx"
OUT = "../Employee_Status.xlsx"


def norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())


def toks(s):
    s = re.sub(r'\b(mr|ms|mrs|dr)\b', '', str(s).lower())
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    return set(t for t in s.split() if len(t) >= 3)


# Known client-side / external people (not Finovate staff) — seen in shared projects
EXTERNAL = {norm(x) for x in [
    "Susan Chalker", "Allison Rinehimer", "Carizza Bernardo", "Michael Hardyman", "Matthew Hidalgo",
    "Nicholas Short", "Neil Maslansky", "Carrie Bowman", "Sheena Ratliff", "Tommy McDonald", "Bryan Hart",
    "Kari Bolton", "Marvin Esquivel", "MARK DI ZAZZO", "Daniel Gregory", "Rachel Galliano", "Jim Kauderer",
    "Daniel Ledbetter", "Phillip Golowatsch", "Daniel Roberts", "Ahnonie Alodia Aguam", "Melissa Knox",
    "Bernadette Bryson", "Christy Nicholas", "Juliet King", "April Hayden", "Kate Carilo", "Trudi Walde",
    "Marcia Vargas Rojas", "Matthew Howard", "Michael Molina", "Peter Tournis", "Ryan Davis", "Arthur Donovan",
    "Cassie Denzler", "Erica Stuart", "test tes",
]}

job = pd.read_excel(JOB)
rel = pd.read_excel(REL)
pool = []  # (norm, tokenset, status, rowdict)
for _, r in job.iterrows():
    pool.append((norm(r["Full Name"]), toks(r["Full Name"]), "ACTIVE", r))
for _, r in rel.iterrows():
    pool.append((norm(r["Full Name"]), toks(r["Full Name"]), "RELIEVED", r))
pool_norms = [p[0] for p in pool]


def match(nm):
    n = norm(nm); tk = toks(nm)
    for p in pool:
        if n == p[0]:
            return p
    fz = difflib.get_close_matches(n, pool_norms, n=1, cutoff=0.86)
    if fz:
        return next(p for p in pool if p[0] == fz[0])
    best = None; bs = 0
    for p in pool:
        ov = len(tk & p[1])
        if ov > bs:
            bs = ov; best = p
    if bs >= 2:
        return best
    return None


hs = db.q("""SELECT user_name nm, round(sum(tracked)/3600.0,1) h, max(date)::text last,
                    round(sum(CASE WHEN coalesce(billable,0)>0 THEN tracked ELSE 0 END)/3600.0,1) bill
             FROM hubstaff_activities WHERE coalesce(user_name,'')<>'' AND coalesce(tracked,0)>0
             GROUP BY 1""")

active, relieved, external, unknown = [], [], [], []
for _, row in hs.iterrows():
    nm, h, last = row["nm"], float(row["h"]), row["last"]
    if norm(nm) in EXTERNAL:
        external.append({"Dashboard Name": nm, "Hours": h, "Last worked": last, "Note": "Client-side / not Finovate staff"})
        continue
    m = match(nm)
    if m is None:
        unknown.append({"Dashboard Name": nm, "Hours": h, "Last worked": last,
                        "Note": "Not in HR roster (left pre-2025-26? / name variant? / missing)"})
    elif m[3] == "ACTIVE" if False else m[2] == "ACTIVE":
        r = m[3]
        active.append({"Dashboard Name": nm, "HR Name": r["Full Name"], "Emp #": r["Employee Number"],
                       "Job Title": r["Job Title"], "Department": r["Department"], "Team": r["Team Name"],
                       "Reporting To": r.get("Reporting To", ""), "Hours": h, "Last worked": last})
    else:
        r = m[3]
        relieved.append({"Dashboard Name": nm, "HR Name": r["Full Name"], "Emp #": r["Employee Number"],
                         "Job Title": r["Job Title"], "Department": r["Department"], "Team": r["Team"],
                         "Exit Date": str(r["Exit Date"])[:10], "Hours": h, "Last worked": last})

dfA = pd.DataFrame(active).sort_values("Hours", ascending=False)
dfR = pd.DataFrame(relieved).sort_values("Hours", ascending=False)
dfE = pd.DataFrame(external).sort_values("Hours", ascending=False)
dfU = pd.DataFrame(unknown).sort_values("Hours", ascending=False)
summ = pd.DataFrame([
    ["ACTIVE (in Job Details)", len(dfA), round(dfA["Hours"].sum(), 0) if len(dfA) else 0],
    ["RELIEVED (left company)", len(dfR), round(dfR["Hours"].sum(), 0) if len(dfR) else 0],
    ["EXTERNAL / CLIENT (not staff)", len(dfE), round(dfE["Hours"].sum(), 0) if len(dfE) else 0],
    ["UNKNOWN-STAFF (not in HR roster)", len(dfU), round(dfU["Hours"].sum(), 0) if len(dfU) else 0],
    ["TOTAL dashboard employees", len(hs), round(float(hs["h"].sum()), 0)],
], columns=["Category", "Count", "Total Hours"])

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    summ.to_excel(xw, sheet_name="Summary", index=False)
    dfA.to_excel(xw, sheet_name="Active", index=False)
    dfR.to_excel(xw, sheet_name="Relieved", index=False)
    dfE.to_excel(xw, sheet_name="External-Client", index=False)
    dfU.to_excel(xw, sheet_name="Unknown-Staff", index=False)
    wb = xw.book
    hf = PatternFill("solid", fgColor="0F172A"); hfont = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="E2E8F0")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for c in ws[1]:
            c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="left", vertical="center")
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 11), 44)
        for i, rrow in enumerate(ws.iter_rows(min_row=2), start=2):
            for c in rrow:
                c.border = Border(bottom=thin)
                if i % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="F8FAFC")

print("WROTE", OUT)
print(summ.to_string(index=False))
